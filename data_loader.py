"""数据读取与参数解析"""
import csv
import bisect
from openpyxl import load_workbook


def load_xlsx(filepath):
    """读取 xlsx 文件第 6 个 sheet 的溶氧/温度/气压数据。

    Returns:
        dict: {timestamps, time_seconds, oxygen, temperature, pressure, nrows}
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[5]]

        timestamps = []
        oxygen = []
        temperature = []
        pressure = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[1] is None or row[6] is None:
                continue
            timestamps.append(row[1])
            oxygen.append(float(row[6]))
            temperature.append(float(row[8]))
            pressure.append(float(row[10]))

        if not timestamps:
            raise ValueError('xlsx contains no data rows (sheet 6 is empty)')

        base = timestamps[0]
        time_seconds = [(t - base).total_seconds() for t in timestamps]

        return {
            'timestamps': timestamps,
            'time_seconds': time_seconds,
            'oxygen': oxygen,
            'temperature': temperature,
            'pressure': pressure,
            'nrows': len(timestamps),
        }
    finally:
        wb.close()


def load_params(csv_path):
    """读取 meas_params.csv。

    Args:
        csv_path: CSV 文件路径

    Returns:
        list[dict]: 每行一个参数字典
    """
    params_list = []
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if not row.get('meas_time', '').strip():
                continue
            p = {
                'meas_time': row['meas_time'].strip(),
                'meas_type': row.get('meas_type', '').strip(),
                'rmr_type': row.get('rmr_type', '').strip(),
                'chamber_id': row.get('chamber_ID', '').strip(),
                'meas_batch': row.get('meas_batch', '').strip(),
                'temperature': row.get('temperature', '').strip(),
                'cycles': int(row.get('cycles', 0) or 0),
                'initial': int(row.get('initial', 0) or 0),
                'cycle_length': int(row.get('cycle_length', 0) or 0),
                'cycle_start': int(row.get('cycle_start', 0) or 0),
                'cycle_time': int(row.get('cycle_time', 0) or 0),
                'flush_time': int(row.get('flush_time', 0) or 0),
                'all_time': int(row.get('all_time', 0) or 0),
            }
            params_list.append(p)
    return params_list


def compute_cycle_boundaries(time_seconds, params):
    """根据循环参数计算每个循环在 time_seconds 数组中的索引边界。

    cycle_time 是斜率窗口终点 (距周期起点的秒数)，不是窗口时长。

    Returns:
        list[dict]: cycle_num, start_idx, end_idx,
            slope_start_idx, slope_end_idx,
            start_seconds, cycle_length,
            slope_start_seconds, slope_end_seconds
    """
    cycles = params['cycles']
    cycle_length = params['cycle_length']
    slope_offset = params['cycle_start']   # 斜率窗口起点 (距周期起点的秒数)
    slope_end = params['cycle_time']       # 斜率窗口终点 (距周期起点的秒数, 非时长!)

    n = len(time_seconds)
    bounds = []

    for i in range(cycles):
        cyc_start_s = i * cycle_length
        cyc_end_s = (i + 1) * cycle_length
        slp_start_s = cyc_start_s + slope_offset
        slp_end_s = cyc_start_s + slope_end     # ← 关键: cycle_time 是终点, 非时长

        a = bisect.bisect_left(time_seconds, cyc_start_s)
        b = bisect.bisect_left(time_seconds, cyc_end_s)
        sa = bisect.bisect_left(time_seconds, slp_start_s)
        sb = bisect.bisect_left(time_seconds, slp_end_s)

        if a >= n:
            break

        bounds.append({
            'cycle_num': i + 1,
            'start_idx': a,
            'end_idx': min(b, n),
            'slope_start_idx': sa,
            'slope_end_idx': min(sb, n),
            'start_seconds': cyc_start_s,
            'cycle_length': cycle_length,
            'slope_start_seconds': slp_start_s,
            'slope_end_seconds': slp_end_s,
        })

    return bounds
